#!/usr/bin/env python3
"""
生成带格式的 Excel 实验结果汇总表。

从 summarize_results_csv.py 的输出 CSV 读取数据，
生成带有条件格式、颜色、加粗等的 xlsx 文件。

Usage:
    conda run -n evalscope python eval/scripts/generate_results_xlsx.py \
        --input /tmp/data_analyst_results.csv \
        --output eval/results/results_summary.xlsx
"""

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── 目录名 → 正式名称映射 ─────────────────────────────────────────────────────

FORMAL_NAMES = {
    # Baselines
    "qwen3_4b_instruct_2507": "Qwen3-4B-Instruct (Baseline)",
    "qwen3_4b_instruct_140k_2phase": "Qwen3-4B-Instruct Baseline naive_mix_140k_2phase",
    "qwen3_4b_instruct_superior_65k_27k_2phase": "Qwen3-4B-Instruct Baseline superior_65k+27k_2phase",
    # Qwen3-4B-Instruct PL-MoE
    "qwen3_4b_instruct_pl_moe_20k": "Qwen3-4B-Instruct PL-MoE naive_mix_20k",
    "qwen3_4b_instruct_pl_moe_20k_a100": "Qwen3-4B-Instruct PL-MoE naive_mix_20k (A100)",
    "qwen3_4b_instruct_pl_moe_140k": "Qwen3-4B-Instruct PL-MoE naive_mix_140k",
    "qwen3_4b_instruct_pl_moe_superior_27k_27k": "Qwen3-4B-Instruct PL-MoE superior_reasoning_27k+27k",
    "qwen3_4b_instruct_pl_moe_superior_65k_27k": "Qwen3-4B-Instruct PL-MoE superior_reasoning_65k+27k",
    # Qwen3-4B-Base PL-MoE
    "qwen3_4b_base_pl_moe_20k": "Qwen3-4B-Base PL-MoE naive_mix_20k",
    "qwen3_4b_base_pl_moe_140k": "Qwen3-4B-Base PL-MoE naive_mix_140k",
    "qwen3_4b_base_pl_moe_superior_27k_27k": "Qwen3-4B-Base PL-MoE superior_reasoning_27k+27k",
    "qwen3_4b_base_pl_moe_superior_65k_27k": "Qwen3-4B-Base PL-MoE superior_reasoning_65k+27k",
    # Qwen3-4B-Base→Instruct Stage2 PL-MoE
    "qwen3_4b_base_instruct_stage2_pl_moe_superior_65k_27k": "Qwen3-4B-Base→Instruct Stage2 PL-MoE superior_reasoning_65k+27k",
    # Qwen3-4B (原始) PL-MoE
    "qwen3_4b_pl_moe_20k": "Qwen3-4B PL-MoE naive_mix_20k",
    "qwen3_4b_pl_moe_superior_27k_27k": "Qwen3-4B PL-MoE superior_reasoning_27k+27k",
    "qwen3_4b_pl_moe_superior_27k_27k_v2": "Qwen3-4B PL-MoE superior_reasoning_27k+27k_v2",
    "qwen3_4b_pl_moe_superior_27k_27k_ckpt2400": "Qwen3-4B PL-MoE superior_reasoning_27k+27k (ckpt2400)",
    "qwen3_4b_pl_moe_superior_27k_27k_aime24_64k": "Qwen3-4B PL-MoE superior_reasoning_27k+27k (AIME24_64k)",
    # Qwen2.5-7B-Instruct PL-MoE
    "qwen_pl_moe_20k": "Qwen2.5-7B-Instruct PL-MoE naive_mix_20k",
    "qwen_pl_moe_140k": "Qwen2.5-7B-Instruct PL-MoE naive_mix_140k",
    # Qwen2.5-7B-Instruct PL-MoE
    "qwen2_pl_moe_superior_27k_27k": "Qwen2.5-7B-Instruct PL-MoE superior_reasoning_27k+27k",
    "qwen2_pl_moe_stage2_superior_27k_27k": "Qwen2.5-7B-Instruct PL-MoE Stage2_superior_reasoning_27k+27k",
    # LLaMA-3-8B PL-MoE
    "llama3_8b_pl_moe_20k": "LLaMA-3-8B-Instruct PL-MoE naive_mix_20k",
    "llama3_8b_pl_moe_superior_27k_27k": "LLaMA-3-8B-Instruct PL-MoE superior_reasoning_27k+27k",
    # Phi-4-Mini PL-MoE
    "phi4_mini_pl_moe_superior_27k_27k": "Phi-4-Mini PL-MoE superior_reasoning_27k+27k",
}

# 模型分组和排序
MODEL_GROUP_ORDER = [
    # (组名, [目录名列表])
    ("Baseline", [
        "qwen3_4b_instruct_2507",
        "qwen3_4b_instruct_140k_2phase",
        "qwen3_4b_instruct_superior_65k_27k_2phase",
    ]),
    ("Qwen3-4B-Instruct PL-MoE (naive-mix)", [
        "qwen3_4b_instruct_pl_moe_20k",
        "qwen3_4b_instruct_pl_moe_20k_a100",
        "qwen3_4b_instruct_pl_moe_140k",
    ]),
    ("Qwen3-4B-Instruct PL-MoE (superior-reasoning)", [
        "qwen3_4b_instruct_pl_moe_superior_27k_27k",
        "qwen3_4b_instruct_pl_moe_superior_65k_27k",
    ]),
    ("Qwen3-4B-Base PL-MoE", [
        "qwen3_4b_base_pl_moe_20k",
        "qwen3_4b_base_pl_moe_140k",
        "qwen3_4b_base_pl_moe_superior_27k_27k",
        "qwen3_4b_base_pl_moe_superior_65k_27k",
        "qwen3_4b_base_instruct_stage2_pl_moe_superior_65k_27k",
    ]),
    ("Qwen3-4B PL-MoE", [
        "qwen3_4b_pl_moe_20k",
        "qwen3_4b_pl_moe_superior_27k_27k",
        "qwen3_4b_pl_moe_superior_27k_27k_v2",
        "qwen3_4b_pl_moe_superior_27k_27k_ckpt2400",
        "qwen3_4b_pl_moe_superior_27k_27k_aime24_64k",
    ]),
    ("Qwen2.5-7B-Instruct PL-MoE", [
        "qwen_pl_moe_20k",
        "qwen_pl_moe_140k",
        "qwen2_pl_moe_superior_27k_27k",
        "qwen2_pl_moe_stage2_superior_27k_27k",
    ]),
    ("LLaMA-3-8B-Instruct PL-MoE", [
        "llama3_8b_pl_moe_20k",
        "llama3_8b_pl_moe_superior_27k_27k",
    ]),
    ("Phi-4-Mini PL-MoE", [
        "phi4_mini_pl_moe_superior_27k_27k",
    ]),
]

BENCHMARKS = ["math500", "aime24", "mmlu_stem", "gpqa_diamond"]
BENCHMARK_DISPLAY = {
    "math500": "MATH500",
    "aime24": "AIME24",
    "mmlu_stem": "MMLU-STEM",
    "gpqa_diamond": "GPQA-Diamond",
}
METRICS = ["acc", "avg_len", "#reflective", "#wait"]
METRIC_DISPLAY = {
    "acc": "Accuracy",
    "avg_len": "Avg Length",
    "#reflective": "Avg #Reflective",
    "#wait": "Avg #Wait",
}

# 分析区域样式
ANALYSIS_TITLE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ANALYSIS_TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
ANALYSIS_BODY_FONT = Font(name="Calibri", size=10)
ANALYSIS_BODY_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
ANALYSIS_BULLET_FONT = Font(name="Calibri", bold=True, size=10, color="2F5496")

# ── 样式定义 ──────────────────────────────────────────────────────────────────

# 颜色
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10)

GROUP_HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GROUP_HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="375623")

BASELINE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

THINK_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
NO_THINK_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

BEST_FONT = Font(name="Calibri", bold=True, color="006100", size=10)
BEST_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

WARN_FONT = Font(name="Calibri", color="9C0006", size=10)
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

NORMAL_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
GROUP_BOTTOM_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="medium", color="8DB4E2"),
)


def write_analysis_section(ws, start_row: int, title: str, lines: list[str], total_cols: int):
    """
    在 sheet 的 start_row 位置写入分析区域。
    title: 分析标题
    lines: 分析条目列表，每条一行
    total_cols: 合并单元格的总列数
    """
    row = start_row + 1  # 空一行

    # 标题行
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = ANALYSIS_TITLE_FONT
    cell.fill = ANALYSIS_TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, total_cols + 1):
        ws.cell(row=row, column=c).fill = ANALYSIS_TITLE_FILL
    row += 1

    # 内容行
    for line in lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        cell = ws.cell(row=row, column=1, value=line)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if line.startswith("▸"):
            cell.font = ANALYSIS_BULLET_FONT
        else:
            cell.font = ANALYSIS_BODY_FONT
        cell.fill = ANALYSIS_BODY_FILL
        for c in range(2, total_cols + 1):
            ws.cell(row=row, column=c).fill = ANALYSIS_BODY_FILL
        ws.row_dimensions[row].height = 20
        row += 1

    return row




def read_csv_data(csv_path: str) -> tuple[dict, dict]:
    """读取 CSV，返回 (data, source_dirs)。
    data: {(dir_name, mode): {(benchmark, metric): value}}
    source_dirs: {(dir_name, mode): source_dir_path}
    """
    data = {}
    source_dirs = {}
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            model = row["model"]
            mode = row["mode"]
            key = (model, mode)
            metrics = {}
            for bench in BENCHMARKS:
                for metric in METRICS:
                    col = f"{bench}_{metric}"
                    val = row.get(col, "")
                    if val:
                        try:
                            metrics[(bench, metric)] = float(val)
                        except ValueError:
                            pass
            data[key] = metrics
            source_dirs[key] = row.get("source_dir", "")
    return data, source_dirs


def find_best_values(data: dict) -> dict:
    """
    找出每个 (benchmark, mode) 下 accuracy 的最大值。
    只在 PL-MoE 模型间比较 (排除 baseline)。
    返回 {(bench, mode): best_acc}
    """
    best = {}
    for (model, mode), metrics in data.items():
        # 排除 baseline
        if "2507" in model or mode == "-":
            continue
        for bench in BENCHMARKS:
            acc = metrics.get((bench, "acc"))
            if acc is not None:
                key = (bench, mode)
                if key not in best or acc > best[key]:
                    best[key] = acc
    return best


def find_length_warnings(data: dict) -> set:
    """
    标记 no_think 模式下 avg_len 异常偏长的单元格。
    条件: no_think avg_len > 3000 (可能存在 reasoning leakage)
    """
    warnings = set()
    for (model, mode), metrics in data.items():
        if mode != "no_think":
            continue
        for bench in BENCHMARKS:
            avg_len = metrics.get((bench, "avg_len"))
            if avg_len is not None and avg_len > 3000:
                warnings.add((model, bench))
    return warnings


def create_accuracy_sheet(wb: Workbook, data: dict, best_values: dict, length_warnings: set,
                          source_dirs: dict):
    """创建 Accuracy 主表 sheet"""
    ws = wb.active
    ws.title = "Accuracy"

    src_col = 3 + len(BENCHMARKS)  # source_dir 列号
    total_cols = src_col

    # ── 表头 ──
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    ws.cell(row=row, column=1, value="Model").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row + 1, column=1).fill = HEADER_FILL

    ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=2)
    ws.cell(row=row, column=2, value="Mode").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row + 1, column=2).fill = HEADER_FILL

    col = 3
    for bench in BENCHMARKS:
        ws.cell(row=row, column=col, value=BENCHMARK_DISPLAY[bench]).font = HEADER_FONT
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        # 子表头
        ws.cell(row=row + 1, column=col, value="Accuracy").font = SUBHEADER_FONT
        ws.cell(row=row + 1, column=col).fill = SUBHEADER_FILL
        ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal="center")
        col += 1

    # Source Dir 表头
    ws.merge_cells(start_row=row, start_column=src_col, end_row=row + 1, end_column=src_col)
    ws.cell(row=row, column=src_col, value="Source Dir").font = HEADER_FONT
    ws.cell(row=row, column=src_col).fill = HEADER_FILL
    ws.cell(row=row, column=src_col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row + 1, column=src_col).fill = HEADER_FILL

    row = 3  # 数据起始行

    for group_name, model_list in MODEL_GROUP_ORDER:
        # 检查组内是否有数据
        group_has_data = False
        for model_dir in model_list:
            for mode in ["think", "no_think", "-"]:
                if (model_dir, mode) in data:
                    group_has_data = True
                    break
            if group_has_data:
                break
        if not group_has_data:
            continue

        # 组标题行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        ws.cell(row=row, column=1, value=group_name).font = GROUP_HEADER_FONT
        ws.cell(row=row, column=1).fill = GROUP_HEADER_FILL
        for c in range(1, total_cols + 1):
            ws.cell(row=row, column=c).fill = GROUP_HEADER_FILL
        row += 1

        for model_dir in model_list:
            # 确定该模型有哪些 mode
            modes_present = []
            for mode in ["think", "no_think", "-"]:
                if (model_dir, mode) in data:
                    modes_present.append(mode)
            if not modes_present:
                continue

            formal_name = FORMAL_NAMES.get(model_dir, model_dir)
            src_written = False  # 每个模型只在第一行写 source_dir

            for mode in modes_present:
                metrics = data[(model_dir, mode)]
                mode_display = mode if mode != "-" else "default"

                is_baseline = (mode == "-")
                bg_fill = BASELINE_FILL if is_baseline else (
                    THINK_FILL if mode == "think" else NO_THINK_FILL
                )

                ws.cell(row=row, column=1, value=formal_name).font = NORMAL_FONT
                ws.cell(row=row, column=1).fill = bg_fill
                ws.cell(row=row, column=2, value=mode_display).font = NORMAL_FONT
                ws.cell(row=row, column=2).fill = bg_fill
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

                col = 3
                for bench in BENCHMARKS:
                    acc = metrics.get((bench, "acc"))
                    cell = ws.cell(row=row, column=col)
                    cell.fill = bg_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = "0.000"

                    if acc is not None:
                        cell.value = acc
                        cell.font = NORMAL_FONT

                        # 高亮 best (PL-MoE 中最好的)
                        if not is_baseline and (bench, mode) in best_values:
                            if acc == best_values[(bench, mode)]:
                                cell.font = BEST_FONT
                                cell.fill = BEST_FILL
                    else:
                        cell.value = "—"
                        cell.font = Font(name="Calibri", size=10, color="AAAAAA")

                    col += 1

                # Source Dir (只在模型第一行写，第二行留空)
                src_cell = ws.cell(row=row, column=src_col)
                src_cell.fill = bg_fill
                if not src_written:
                    src_cell.value = source_dirs.get((model_dir, mode), "")
                    src_cell.font = Font(name="Calibri", size=8, color="808080")
                    src_cell.alignment = Alignment(horizontal="left", vertical="center")
                    src_written = True

                row += 1

        # 组间分隔: 给最后一行加底部边框
        for c in range(1, total_cols + 1):
            ws.cell(row=row - 1, column=c).border = GROUP_BOTTOM_BORDER

    # ── 列宽 ──
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    for i, bench in enumerate(BENCHMARKS):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16
    ws.column_dimensions[get_column_letter(src_col)].width = 55

    # 冻结窗格
    ws.freeze_panes = "C3"

    return row


def _write_source_dir_header(ws, row, src_col):
    """写 Source Dir 表头 (两行合并)"""
    ws.merge_cells(start_row=row, start_column=src_col, end_row=row + 1, end_column=src_col)
    ws.cell(row=row, column=src_col, value="Source Dir").font = HEADER_FONT
    ws.cell(row=row, column=src_col).fill = HEADER_FILL
    ws.cell(row=row, column=src_col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row + 1, column=src_col).fill = HEADER_FILL


def _write_source_dir_cell(ws, row, src_col, source_dirs, model_dir, mode, bg_fill, src_written):
    """写数据行的 Source Dir 单元格。返回更新后的 src_written。"""
    src_cell = ws.cell(row=row, column=src_col)
    src_cell.fill = bg_fill
    if not src_written:
        src_cell.value = source_dirs.get((model_dir, mode), "")
        src_cell.font = Font(name="Calibri", size=8, color="808080")
        src_cell.alignment = Alignment(horizontal="left", vertical="center")
        return True
    return src_written


def create_length_sheet(wb: Workbook, data: dict, length_warnings: set, source_dirs: dict):
    """创建 Output Length sheet"""
    ws = wb.create_sheet(title="Output Length")

    src_col = 3 + len(BENCHMARKS)
    total_cols = src_col

    # ── 表头 ──
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    ws.cell(row=row, column=1, value="Model").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row + 1, column=1).fill = HEADER_FILL

    ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=2)
    ws.cell(row=row, column=2, value="Mode").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row + 1, column=2).fill = HEADER_FILL

    col = 3
    for bench in BENCHMARKS:
        ws.cell(row=row, column=col, value=BENCHMARK_DISPLAY[bench]).font = HEADER_FONT
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=row + 1, column=col, value="Avg Length").font = SUBHEADER_FONT
        ws.cell(row=row + 1, column=col).fill = SUBHEADER_FILL
        ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal="center")
        col += 1

    _write_source_dir_header(ws, row, src_col)

    row = 3

    for group_name, model_list in MODEL_GROUP_ORDER:
        group_has_data = any(
            (m, mode) in data
            for m in model_list for mode in ["think", "no_think", "-"]
        )
        if not group_has_data:
            continue

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        ws.cell(row=row, column=1, value=group_name).font = GROUP_HEADER_FONT
        ws.cell(row=row, column=1).fill = GROUP_HEADER_FILL
        for c in range(1, total_cols + 1):
            ws.cell(row=row, column=c).fill = GROUP_HEADER_FILL
        row += 1

        for model_dir in model_list:
            modes_present = [m for m in ["think", "no_think", "-"] if (model_dir, m) in data]
            if not modes_present:
                continue

            formal_name = FORMAL_NAMES.get(model_dir, model_dir)
            src_written = False

            for mode in modes_present:
                metrics = data[(model_dir, mode)]
                mode_display = mode if mode != "-" else "default"
                is_baseline = (mode == "-")
                bg_fill = BASELINE_FILL if is_baseline else (
                    THINK_FILL if mode == "think" else NO_THINK_FILL
                )

                ws.cell(row=row, column=1, value=formal_name).font = NORMAL_FONT
                ws.cell(row=row, column=1).fill = bg_fill
                ws.cell(row=row, column=2, value=mode_display).font = NORMAL_FONT
                ws.cell(row=row, column=2).fill = bg_fill
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

                col = 3
                for bench in BENCHMARKS:
                    avg_len = metrics.get((bench, "avg_len"))
                    cell = ws.cell(row=row, column=col)
                    cell.fill = bg_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = "#,##0"

                    if avg_len is not None:
                        cell.value = avg_len
                        if mode == "no_think" and (model_dir, bench) in length_warnings:
                            cell.font = WARN_FONT
                            cell.fill = WARN_FILL
                        else:
                            cell.font = NORMAL_FONT
                    else:
                        cell.value = "—"
                        cell.font = Font(name="Calibri", size=10, color="AAAAAA")

                    col += 1

                src_written = _write_source_dir_cell(
                    ws, row, src_col, source_dirs, model_dir, mode, bg_fill, src_written)

                row += 1

        for c in range(1, total_cols + 1):
            ws.cell(row=row - 1, column=c).border = GROUP_BOTTOM_BORDER

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    for i in range(len(BENCHMARKS)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16
    ws.column_dimensions[get_column_letter(src_col)].width = 55
    ws.freeze_panes = "C3"


def create_reflective_sheet(wb: Workbook, data: dict, source_dirs: dict):
    """创建 Reflective Tokens sheet (只看 #reflective 和 #wait)"""
    ws = wb.create_sheet(title="Reflective Tokens")

    src_col = 3 + len(BENCHMARKS) * 2
    total_cols = src_col

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    ws.cell(row=row, column=1, value="Model").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row + 1, column=1).fill = HEADER_FILL

    ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=2)
    ws.cell(row=row, column=2, value="Mode").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row + 1, column=2).fill = HEADER_FILL

    col = 3
    for bench in BENCHMARKS:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.cell(row=row, column=col, value=BENCHMARK_DISPLAY[bench]).font = HEADER_FONT
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=col + 1).fill = HEADER_FILL

        ws.cell(row=row + 1, column=col, value="#Reflective").font = SUBHEADER_FONT
        ws.cell(row=row + 1, column=col).fill = SUBHEADER_FILL
        ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal="center")

        ws.cell(row=row + 1, column=col + 1, value="#Wait").font = SUBHEADER_FONT
        ws.cell(row=row + 1, column=col + 1).fill = SUBHEADER_FILL
        ws.cell(row=row + 1, column=col + 1).alignment = Alignment(horizontal="center")
        col += 2

    _write_source_dir_header(ws, row, src_col)

    row = 3

    for group_name, model_list in MODEL_GROUP_ORDER:
        group_has_data = any(
            (m, mode) in data
            for m in model_list for mode in ["think", "no_think", "-"]
        )
        if not group_has_data:
            continue

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        ws.cell(row=row, column=1, value=group_name).font = GROUP_HEADER_FONT
        ws.cell(row=row, column=1).fill = GROUP_HEADER_FILL
        for c in range(1, total_cols + 1):
            ws.cell(row=row, column=c).fill = GROUP_HEADER_FILL
        row += 1

        for model_dir in model_list:
            modes_present = [m for m in ["think", "no_think", "-"] if (model_dir, m) in data]
            if not modes_present:
                continue

            formal_name = FORMAL_NAMES.get(model_dir, model_dir)
            src_written = False

            for mode in modes_present:
                metrics = data[(model_dir, mode)]
                mode_display = mode if mode != "-" else "default"
                is_baseline = (mode == "-")
                bg_fill = BASELINE_FILL if is_baseline else (
                    THINK_FILL if mode == "think" else NO_THINK_FILL
                )

                ws.cell(row=row, column=1, value=formal_name).font = NORMAL_FONT
                ws.cell(row=row, column=1).fill = bg_fill
                ws.cell(row=row, column=2, value=mode_display).font = NORMAL_FONT
                ws.cell(row=row, column=2).fill = bg_fill
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

                col = 3
                for bench in BENCHMARKS:
                    for metric_key in ["#reflective", "#wait"]:
                        val = metrics.get((bench, metric_key))
                        cell = ws.cell(row=row, column=col)
                        cell.fill = bg_fill
                        cell.alignment = Alignment(horizontal="center")
                        cell.number_format = "0.0000"

                        if val is not None:
                            cell.value = val
                            if mode == "no_think" and val > 0.01:
                                cell.font = WARN_FONT
                                cell.fill = WARN_FILL
                            else:
                                cell.font = NORMAL_FONT
                        else:
                            cell.value = "—"
                            cell.font = Font(name="Calibri", size=10, color="AAAAAA")
                        col += 1

                src_written = _write_source_dir_cell(
                    ws, row, src_col, source_dirs, model_dir, mode, bg_fill, src_written)

                row += 1

        for c in range(1, total_cols + 1):
            ws.cell(row=row - 1, column=c).border = GROUP_BOTTOM_BORDER

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    for i in range(len(BENCHMARKS) * 2):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14
    ws.column_dimensions[get_column_letter(src_col)].width = 55
    ws.freeze_panes = "C3"


def create_legend_sheet(wb: Workbook):
    """创建图例说明 sheet"""
    ws = wb.create_sheet(title="Legend")

    ws.cell(row=1, column=1, value="格式说明").font = Font(name="Calibri", bold=True, size=14)

    items = [
        (3, BASELINE_FILL, NORMAL_FONT, "黄底", "Baseline 模型 (原始模型，未经 PL-MoE 训练)"),
        (4, THINK_FILL, NORMAL_FONT, "白底", "think 模式 (使用 expert 1, CoT 推理)"),
        (5, NO_THINK_FILL, NORMAL_FONT, "灰底", "no_think 模式 (使用 expert 0, 直接回答)"),
        (6, BEST_FILL, BEST_FONT, "绿底加粗", "PL-MoE 模型中该 benchmark + mode 下的最高 accuracy"),
        (7, WARN_FILL, WARN_FONT, "红底", "异常值警告 (no_think 输出过长或出现 reflective tokens)"),
    ]
    for row_num, fill, font, label, desc in items:
        ws.cell(row=row_num, column=1, value=label).fill = fill
        ws.cell(row=row_num, column=1).font = font
        ws.cell(row=row_num, column=2, value=desc).font = NORMAL_FONT

    ws.cell(row=9, column=1, value="Sheet 说明").font = Font(name="Calibri", bold=True, size=14)
    sheets_info = [
        (10, "Accuracy", "各模型在 4 个 benchmark 上的准确率，绿色高亮为同模式下最优"),
        (11, "Output Length", "平均输出 token 数，红色标记 no_think 下 > 3000 tokens 的异常 (可能存在 reasoning leakage)"),
        (12, "Reflective Tokens", "平均 reflective token 数 (#reflective 包含 wait/hmm/okay/alternatively/let me think, #wait 仅统计 'wait, ')"),
    ]
    for row_num, sheet_name, desc in sheets_info:
        ws.cell(row=row_num, column=1, value=sheet_name).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row_num, column=2, value=desc).font = NORMAL_FONT

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80


def main():
    parser = argparse.ArgumentParser(description="生成带格式的 Excel 实验结果汇总表")
    parser.add_argument("--input", type=str, required=True, help="输入 CSV 文件路径")
    parser.add_argument("--output", type=str, required=True, help="输出 xlsx 文件路径")
    args = parser.parse_args()

    # 读取数据
    data, source_dirs = read_csv_data(args.input)
    print(f"读取 {len(data)} 行数据")

    # 计算高亮信息
    best_values = find_best_values(data)
    length_warnings = find_length_warnings(data)

    # 创建工作簿
    wb = Workbook()
    create_accuracy_sheet(wb, data, best_values, length_warnings, source_dirs)
    create_length_sheet(wb, data, length_warnings, source_dirs)
    create_reflective_sheet(wb, data, source_dirs)
    create_legend_sheet(wb)

    # 保存
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"已生成: {output_path}")


if __name__ == "__main__":
    main()
